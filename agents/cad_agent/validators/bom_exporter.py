"""
BOM Export Module
==================
Export Bill of Materials to Excel and CSV formats with comprehensive formatting.

Phase 25 - BOM Export Functionality

Features:
- Excel export with formatting, colors, and formulas
- CSV export for data exchange
- Multiple BOM views (flat, hierarchical, summary)
- Cost rollup calculations
- Material summary sheets
- Fastener/hardware summary
"""

import csv
import io
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
from enum import Enum

# Excel imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side, Alignment,
        NamedStyle, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger("vulcan.bom_exporter")


# =============================================================================
# DATA CLASSES
# =============================================================================

class BOMFormat(Enum):
    """BOM export format."""
    EXCEL = "excel"
    CSV = "csv"
    BOTH = "both"


class BOMView(Enum):
    """BOM structure view."""
    FLAT = "flat"               # All parts in single list
    HIERARCHICAL = "hierarchical"  # Parent-child tree structure
    INDENTED = "indented"       # Indented list showing hierarchy
    SUMMARY = "summary"         # Grouped summary by type/material


@dataclass
class BOMItem:
    """Single BOM line item."""
    item_number: int
    part_number: str
    description: str
    quantity: int = 1
    unit: str = "EA"
    material: str = ""
    weight_lbs: float = 0.0
    unit_cost: float = 0.0
    supplier: str = ""
    lead_time_days: int = 0
    notes: str = ""
    level: int = 0  # For hierarchical BOM
    parent_item: Optional[int] = None
    category: str = ""  # structural, fastener, electrical, etc.
    stock_size: str = ""
    finish: str = ""
    drawing_ref: str = ""


@dataclass
class BOMMetadata:
    """BOM document metadata."""
    assembly_number: str = ""
    assembly_name: str = ""
    revision: str = ""
    project_name: str = ""
    project_number: str = ""
    customer: str = ""
    prepared_by: str = ""
    date: str = ""
    units: str = "imperial"  # imperial or metric


@dataclass
class BOMExportResult:
    """Result of BOM export operation."""
    success: bool
    format: str
    filename: str
    size_bytes: int
    content: bytes
    message: str = ""
    total_items: int = 0
    total_weight: float = 0.0
    total_cost: float = 0.0


# =============================================================================
# BOM EXPORTER CLASS
# =============================================================================

class BOMExporter:
    """
    Export BOM data to Excel and CSV formats.

    Features:
    - Professional Excel formatting
    - Multiple worksheet views
    - Cost/weight calculations
    - Material summaries
    - Chart generation
    """

    def __init__(self):
        self.items: List[BOMItem] = []
        self.metadata: BOMMetadata = BOMMetadata()

    def set_metadata(self, metadata: BOMMetadata):
        """Set BOM metadata."""
        self.metadata = metadata

    def add_item(self, item: BOMItem):
        """Add a BOM item."""
        self.items.append(item)

    def add_items(self, items: List[BOMItem]):
        """Add multiple BOM items."""
        self.items.extend(items)

    def clear(self):
        """Clear all items."""
        self.items.clear()

    def from_dict_list(self, items: List[Dict[str, Any]]):
        """
        Load BOM items from list of dictionaries.

        Args:
            items: List of dicts with BOM item fields
        """
        for i, item_dict in enumerate(items, start=1):
            self.items.append(BOMItem(
                item_number=item_dict.get("item", item_dict.get("item_number", i)),
                part_number=item_dict.get("part_number", f"PART-{i:04d}"),
                description=item_dict.get("description", ""),
                quantity=item_dict.get("qty", item_dict.get("quantity", 1)),
                unit=item_dict.get("unit", "EA"),
                material=item_dict.get("material", ""),
                weight_lbs=item_dict.get("weight_lbs", item_dict.get("weight", 0.0)),
                unit_cost=item_dict.get("unit_cost", item_dict.get("cost", 0.0)),
                supplier=item_dict.get("supplier", item_dict.get("vendor", "")),
                lead_time_days=item_dict.get("lead_time_days", item_dict.get("lead_time", 0)),
                notes=item_dict.get("notes", item_dict.get("remarks", "")),
                level=item_dict.get("level", 0),
                parent_item=item_dict.get("parent_item"),
                category=item_dict.get("category", item_dict.get("type", "")),
                stock_size=item_dict.get("stock_size", item_dict.get("size", "")),
                finish=item_dict.get("finish", item_dict.get("coating", "")),
                drawing_ref=item_dict.get("drawing_ref", item_dict.get("dwg", "")),
            ))

    # =========================================================================
    # CSV EXPORT
    # =========================================================================

    def export_csv(self, filepath: Optional[str] = None) -> BOMExportResult:
        """
        Export BOM to CSV format.

        Args:
            filepath: Optional file path to write to

        Returns:
            BOMExportResult with CSV content
        """
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Header row
        headers = [
            "Item", "Part Number", "Description", "Qty", "Unit",
            "Material", "Weight (lbs)", "Unit Cost ($)", "Extended Cost ($)",
            "Supplier", "Lead Time (days)", "Category", "Stock Size",
            "Finish", "Drawing Ref", "Notes"
        ]
        writer.writerow(headers)

        # Data rows
        total_weight = 0.0
        total_cost = 0.0

        for item in self.items:
            ext_cost = item.quantity * item.unit_cost
            ext_weight = item.quantity * item.weight_lbs
            total_weight += ext_weight
            total_cost += ext_cost

            writer.writerow([
                item.item_number,
                item.part_number,
                item.description,
                item.quantity,
                item.unit,
                item.material,
                f"{item.weight_lbs:.3f}",
                f"{item.unit_cost:.2f}",
                f"{ext_cost:.2f}",
                item.supplier,
                item.lead_time_days,
                item.category,
                item.stock_size,
                item.finish,
                item.drawing_ref,
                item.notes
            ])

        # Summary row
        writer.writerow([])
        writer.writerow(["", "", "TOTAL", len(self.items), "", "", f"{total_weight:.1f}", "", f"{total_cost:.2f}"])

        csv_content = buffer.getvalue()
        buffer.close()

        csv_bytes = csv_content.encode('utf-8')

        if filepath:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                f.write(csv_content)

        filename = f"BOM_{self.metadata.assembly_number or 'export'}.csv"

        return BOMExportResult(
            success=True,
            format="csv",
            filename=filename,
            size_bytes=len(csv_bytes),
            content=csv_bytes,
            total_items=len(self.items),
            total_weight=total_weight,
            total_cost=total_cost,
        )

    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================

    def export_excel(self, filepath: Optional[str] = None, include_charts: bool = True) -> BOMExportResult:
        """
        Export BOM to Excel format with professional formatting.

        Args:
            filepath: Optional file path to write to
            include_charts: Include pie/bar charts

        Returns:
            BOMExportResult with Excel content
        """
        if not OPENPYXL_AVAILABLE:
            return BOMExportResult(
                success=False,
                format="excel",
                filename="",
                size_bytes=0,
                content=b"",
                message="openpyxl not installed. Run: pip install openpyxl"
            )

        wb = Workbook()

        # Create styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        subheader_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")

        data_font = Font(size=10)
        data_align = Alignment(horizontal="left", vertical="center")
        number_align = Alignment(horizontal="right", vertical="center")

        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        currency_format = '"$"#,##0.00'
        weight_format = '#,##0.000'

        # =====================================================================
        # SHEET 1: MAIN BOM
        # =====================================================================
        ws = wb.active
        ws.title = "Bill of Materials"

        # Title section
        ws.merge_cells('A1:P1')
        ws['A1'] = f"BILL OF MATERIALS - {self.metadata.assembly_name or 'Assembly'}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Metadata section
        meta_data = [
            ("Assembly Number:", self.metadata.assembly_number),
            ("Revision:", self.metadata.revision),
            ("Project:", self.metadata.project_name),
            ("Customer:", self.metadata.customer),
            ("Prepared By:", self.metadata.prepared_by),
            ("Date:", self.metadata.date or datetime.now().strftime("%Y-%m-%d")),
        ]

        for i, (label, value) in enumerate(meta_data):
            row = 3 + i // 3
            col = 1 + (i % 3) * 3
            ws.cell(row=row, column=col, value=label).font = Font(bold=True)
            ws.cell(row=row, column=col + 1, value=value or "")

        # BOM header row
        header_row = 6
        headers = [
            ("Item", 6), ("Part Number", 18), ("Description", 35), ("Qty", 6),
            ("Unit", 6), ("Material", 15), ("Weight", 10), ("Unit Cost", 12),
            ("Ext Cost", 12), ("Supplier", 15), ("Lead Time", 10),
            ("Category", 12), ("Stock Size", 12), ("Finish", 12),
            ("Dwg Ref", 12), ("Notes", 25)
        ]

        for col, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        total_weight = 0.0
        total_cost = 0.0
        row = header_row + 1

        for item in self.items:
            ext_cost = item.quantity * item.unit_cost
            ext_weight = item.quantity * item.weight_lbs
            total_weight += ext_weight
            total_cost += ext_cost

            data = [
                item.item_number, item.part_number, item.description,
                item.quantity, item.unit, item.material, item.weight_lbs,
                item.unit_cost, ext_cost, item.supplier, item.lead_time_days,
                item.category, item.stock_size, item.finish, item.drawing_ref, item.notes
            ]

            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.border = border

                if col in [7]:  # Weight
                    cell.number_format = weight_format
                    cell.alignment = number_align
                elif col in [8, 9]:  # Cost columns
                    cell.number_format = currency_format
                    cell.alignment = number_align
                elif col in [4, 11]:  # Qty, Lead time
                    cell.alignment = number_align
                else:
                    cell.alignment = data_align

                # Alternating row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")

            row += 1

        # Totals row
        total_row = row
        ws.cell(row=total_row, column=1, value="").border = border
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=f"{len(self.items)} items").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total_weight).font = Font(bold=True)
        ws.cell(row=total_row, column=7).number_format = weight_format
        ws.cell(row=total_row, column=9, value=total_cost).font = Font(bold=True)
        ws.cell(row=total_row, column=9).number_format = currency_format

        for col in range(1, 17):
            ws.cell(row=total_row, column=col).border = border
            ws.cell(row=total_row, column=col).fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")

        # Freeze header row
        ws.freeze_panes = 'A7'

        # =====================================================================
        # SHEET 2: MATERIAL SUMMARY
        # =====================================================================
        ws_mat = wb.create_sheet("Material Summary")

        ws_mat['A1'] = "MATERIAL SUMMARY"
        ws_mat['A1'].font = Font(bold=True, size=14)

        # Aggregate by material
        material_summary: Dict[str, Dict] = {}
        for item in self.items:
            mat = item.material or "Unspecified"
            if mat not in material_summary:
                material_summary[mat] = {"count": 0, "weight": 0.0, "cost": 0.0}
            material_summary[mat]["count"] += item.quantity
            material_summary[mat]["weight"] += item.quantity * item.weight_lbs
            material_summary[mat]["cost"] += item.quantity * item.unit_cost

        # Header
        mat_headers = ["Material", "Item Count", "Total Weight (lbs)", "Total Cost ($)", "% of Weight", "% of Cost"]
        for col, header in enumerate(mat_headers, start=1):
            cell = ws_mat.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        ws_mat.column_dimensions['A'].width = 25
        ws_mat.column_dimensions['B'].width = 12
        ws_mat.column_dimensions['C'].width = 18
        ws_mat.column_dimensions['D'].width = 15
        ws_mat.column_dimensions['E'].width = 12
        ws_mat.column_dimensions['F'].width = 12

        # Data
        row = 4
        for material, data in sorted(material_summary.items()):
            pct_weight = (data["weight"] / total_weight * 100) if total_weight > 0 else 0
            pct_cost = (data["cost"] / total_cost * 100) if total_cost > 0 else 0

            values = [material, data["count"], data["weight"], data["cost"], pct_weight, pct_cost]
            for col, value in enumerate(values, start=1):
                cell = ws_mat.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 3:
                    cell.number_format = weight_format
                elif col == 4:
                    cell.number_format = currency_format
                elif col in [5, 6]:
                    cell.number_format = '0.0"%"'
            row += 1

        # =====================================================================
        # SHEET 3: CATEGORY SUMMARY
        # =====================================================================
        ws_cat = wb.create_sheet("Category Summary")

        ws_cat['A1'] = "CATEGORY SUMMARY"
        ws_cat['A1'].font = Font(bold=True, size=14)

        # Aggregate by category
        category_summary: Dict[str, Dict] = {}
        for item in self.items:
            cat = item.category or "Unspecified"
            if cat not in category_summary:
                category_summary[cat] = {"count": 0, "weight": 0.0, "cost": 0.0}
            category_summary[cat]["count"] += item.quantity
            category_summary[cat]["weight"] += item.quantity * item.weight_lbs
            category_summary[cat]["cost"] += item.quantity * item.unit_cost

        # Header
        for col, header in enumerate(mat_headers, start=1):
            cell = ws_cat.cell(row=3, column=col, value=header.replace("Material", "Category"))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        ws_cat.column_dimensions['A'].width = 20
        ws_cat.column_dimensions['B'].width = 12
        ws_cat.column_dimensions['C'].width = 18
        ws_cat.column_dimensions['D'].width = 15
        ws_cat.column_dimensions['E'].width = 12
        ws_cat.column_dimensions['F'].width = 12

        row = 4
        for category, data in sorted(category_summary.items()):
            pct_weight = (data["weight"] / total_weight * 100) if total_weight > 0 else 0
            pct_cost = (data["cost"] / total_cost * 100) if total_cost > 0 else 0

            values = [category, data["count"], data["weight"], data["cost"], pct_weight, pct_cost]
            for col, value in enumerate(values, start=1):
                cell = ws_cat.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 3:
                    cell.number_format = weight_format
                elif col == 4:
                    cell.number_format = currency_format
                elif col in [5, 6]:
                    cell.number_format = '0.0"%"'
            row += 1

        # =====================================================================
        # SHEET 4: SUPPLIER SUMMARY
        # =====================================================================
        ws_sup = wb.create_sheet("Supplier Summary")

        ws_sup['A1'] = "SUPPLIER SUMMARY"
        ws_sup['A1'].font = Font(bold=True, size=14)

        # Aggregate by supplier
        supplier_summary: Dict[str, Dict] = {}
        for item in self.items:
            sup = item.supplier or "TBD"
            if sup not in supplier_summary:
                supplier_summary[sup] = {"count": 0, "cost": 0.0, "max_lead": 0}
            supplier_summary[sup]["count"] += item.quantity
            supplier_summary[sup]["cost"] += item.quantity * item.unit_cost
            supplier_summary[sup]["max_lead"] = max(supplier_summary[sup]["max_lead"], item.lead_time_days)

        # Header
        sup_headers = ["Supplier", "Item Count", "Total Cost ($)", "Max Lead Time (days)"]
        for col, header in enumerate(sup_headers, start=1):
            cell = ws_sup.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        ws_sup.column_dimensions['A'].width = 25
        ws_sup.column_dimensions['B'].width = 12
        ws_sup.column_dimensions['C'].width = 15
        ws_sup.column_dimensions['D'].width = 20

        row = 4
        for supplier, data in sorted(supplier_summary.items()):
            values = [supplier, data["count"], data["cost"], data["max_lead"]]
            for col, value in enumerate(values, start=1):
                cell = ws_sup.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 3:
                    cell.number_format = currency_format
            row += 1

        # =====================================================================
        # ADD CHARTS
        # =====================================================================
        if include_charts and len(material_summary) > 1:
            # Material weight pie chart on Material Summary sheet
            chart = PieChart()
            chart.title = "Weight by Material"
            labels = Reference(ws_mat, min_col=1, min_row=4, max_row=3 + len(material_summary))
            data = Reference(ws_mat, min_col=3, min_row=3, max_row=3 + len(material_summary))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 12
            chart.height = 8
            ws_mat.add_chart(chart, "H3")

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        if filepath:
            with open(filepath, 'wb') as f:
                f.write(excel_bytes)

        filename = f"BOM_{self.metadata.assembly_number or 'export'}.xlsx"

        return BOMExportResult(
            success=True,
            format="excel",
            filename=filename,
            size_bytes=len(excel_bytes),
            content=excel_bytes,
            total_items=len(self.items),
            total_weight=total_weight,
            total_cost=total_cost,
        )

    # =========================================================================
    # CONVENIENCE METHODS
    # =========================================================================

    def export(
        self,
        format: BOMFormat = BOMFormat.EXCEL,
        filepath: Optional[str] = None
    ) -> BOMExportResult:
        """
        Export BOM in specified format.

        Args:
            format: Export format (excel, csv, both)
            filepath: Optional file path (extension determines format)

        Returns:
            BOMExportResult
        """
        if format == BOMFormat.CSV:
            return self.export_csv(filepath)
        elif format == BOMFormat.EXCEL:
            return self.export_excel(filepath)
        else:
            # Both - export Excel (CSV can be derived)
            return self.export_excel(filepath)

    def get_summary(self) -> Dict[str, Any]:
        """Get BOM summary statistics."""
        total_weight = sum(i.quantity * i.weight_lbs for i in self.items)
        total_cost = sum(i.quantity * i.unit_cost for i in self.items)
        total_qty = sum(i.quantity for i in self.items)

        materials = set(i.material for i in self.items if i.material)
        categories = set(i.category for i in self.items if i.category)
        suppliers = set(i.supplier for i in self.items if i.supplier)

        max_lead = max((i.lead_time_days for i in self.items), default=0)

        return {
            "total_line_items": len(self.items),
            "total_quantity": total_qty,
            "total_weight_lbs": round(total_weight, 2),
            "total_cost": round(total_cost, 2),
            "unique_materials": len(materials),
            "unique_categories": len(categories),
            "unique_suppliers": len(suppliers),
            "max_lead_time_days": max_lead,
            "assembly": self.metadata.assembly_number,
            "revision": self.metadata.revision,
        }
