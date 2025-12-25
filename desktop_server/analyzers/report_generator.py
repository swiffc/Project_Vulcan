"""
PDF Report Generator
====================
Generate comprehensive PDF reports for ACHE designs.

Phase 24.26 Implementation
"""

import io
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from dataclasses import dataclass

logger = logging.getLogger("vulcan.analyzer.report")


@dataclass
class ReportSection:
    """A section in the report."""
    title: str
    content: List[Dict[str, Any]]
    include: bool = True


class ReportGenerator:
    """
    Generate PDF reports for ACHE designs.
    Uses reportlab for PDF generation.
    """

    def __init__(self):
        self._pdf_available = False
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            self._pdf_available = True
        except ImportError:
            logger.warning("reportlab not installed - PDF generation disabled")

    def _create_header(self, model_data: Dict[str, Any]) -> List:
        """Create report header section."""
        from reportlab.platypus import Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title = model_data.get("document_name", "Model Report")
        elements.append(Paragraph(f"<b>ACHE Design Report</b>", styles["Title"]))
        elements.append(Paragraph(f"Model: {title}", styles["Heading2"]))
        elements.append(Spacer(1, 12))

        # Report info
        job_info = model_data.get("job_info", {})
        info_text = f"""
        <b>Job Number:</b> {job_info.get('job_number', 'N/A')}<br/>
        <b>Customer:</b> {job_info.get('customer', 'N/A')}<br/>
        <b>Tag Number:</b> {job_info.get('tag_number', 'N/A')}<br/>
        <b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>
        """
        elements.append(Paragraph(info_text, styles["Normal"]))
        elements.append(Spacer(1, 24))

        return elements

    def _create_properties_table(self, properties: Dict[str, Any]) -> List:
        """Create properties summary table."""
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("<b>Design Properties</b>", styles["Heading2"]))
        elements.append(Spacer(1, 6))

        # Mass properties
        mass = properties.get("mass", {})
        data = [
            ["Property", "Value", "Units"],
            ["Mass", f"{mass.get('mass_kg', 0):.1f}", "kg"],
            ["Mass", f"{mass.get('mass_lbs', 0):.1f}", "lbs"],
            ["Volume", f"{mass.get('volume_m3', 0):.6f}", "m³"],
            ["Surface Area", f"{mass.get('surface_area_m2', 0):.3f}", "m²"],
        ]

        table = Table(data, colWidths=[150, 100, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))

        return elements

    def _create_compliance_table(self, compliance: Dict[str, Any]) -> List:
        """Create compliance check results table."""
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("<b>Standards Compliance</b>", styles["Heading2"]))
        elements.append(Spacer(1, 6))

        summary = compliance.get("summary", "No checks performed")
        elements.append(Paragraph(f"<i>{summary}</i>", styles["Normal"]))
        elements.append(Spacer(1, 6))

        checks = compliance.get("checks", [])
        if checks:
            data = [["Standard", "Section", "Result", "Notes"]]
            for check in checks[:15]:  # Limit to 15 rows
                result = check.get("level", "n/a").upper()
                data.append([
                    check.get("standard", ""),
                    check.get("section", ""),
                    result,
                    check.get("notes", "")[:30],
                ])

            table = Table(data, colWidths=[80, 80, 60, 150])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(table)

        elements.append(Spacer(1, 12))
        return elements

    def _create_issues_section(self, issues: List[str]) -> List:
        """Create issues/warnings section."""
        from reportlab.platypus import Paragraph, Spacer, ListFlowable, ListItem
        from reportlab.lib.styles import getSampleStyleSheet

        styles = getSampleStyleSheet()
        elements = []

        if issues:
            elements.append(Paragraph("<b>Issues & Warnings</b>", styles["Heading2"]))
            elements.append(Spacer(1, 6))

            for issue in issues[:10]:
                elements.append(Paragraph(f"• {issue}", styles["Normal"]))

            elements.append(Spacer(1, 12))

        return elements

    def generate_pdf(self, model_data: Dict[str, Any]) -> Optional[bytes]:
        """Generate PDF report and return as bytes."""
        if not self._pdf_available:
            logger.error("PDF generation not available - install reportlab")
            return None

        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Spacer

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72,
        )

        elements = []

        # Add sections
        elements.extend(self._create_header(model_data))

        if "properties" in model_data:
            elements.extend(self._create_properties_table(model_data["properties"]))

        if "compliance" in model_data:
            elements.extend(self._create_compliance_table(model_data["compliance"]))

        # Collect all issues
        all_issues = []
        if "interference_check" in model_data:
            all_issues.extend(model_data["interference_check"].get("issues", []))
        if "bend_analysis" in model_data:
            all_issues.extend(model_data["bend_analysis"].get("warnings", []))
        if "mates_analysis" in model_data:
            all_issues.extend(model_data["mates_analysis"].get("issues", []))

        if all_issues:
            elements.extend(self._create_issues_section(all_issues))

        # Build PDF
        try:
            doc.build(elements)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            return pdf_bytes
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            return None

    def generate_excel(self, model_data: Dict[str, Any]) -> Optional[bytes]:
        """Generate Excel report (BOM, properties, checks)."""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            logger.warning("openpyxl not installed - Excel export disabled")
            return None

        wb = Workbook()

        # Properties sheet
        ws_props = wb.active
        ws_props.title = "Properties"
        props = model_data.get("properties", {})

        row = 1
        ws_props.cell(row=row, column=1, value="Property")
        ws_props.cell(row=row, column=2, value="Value")

        # Standard properties
        standard = props.get("standard", {})
        for key, value in standard.items():
            row += 1
            ws_props.cell(row=row, column=1, value=key)
            ws_props.cell(row=row, column=2, value=str(value))

        # Mass properties
        mass = props.get("mass", {})
        for key, value in mass.items():
            row += 1
            ws_props.cell(row=row, column=1, value=key)
            ws_props.cell(row=row, column=2, value=str(value))

        # BOM sheet if available
        if "bom" in model_data:
            ws_bom = wb.create_sheet("BOM")
            bom = model_data["bom"]
            if isinstance(bom, list) and bom:
                headers = list(bom[0].keys())
                for col, header in enumerate(headers, 1):
                    ws_bom.cell(row=1, column=col, value=header)
                for row_num, item in enumerate(bom, 2):
                    for col, header in enumerate(headers, 1):
                        ws_bom.cell(row=row_num, column=col, value=item.get(header, ""))

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def to_dict(self, model_data: Dict[str, Any], format: str = "pdf") -> Dict[str, Any]:
        """Generate report and return metadata."""
        if format == "pdf":
            pdf_bytes = self.generate_pdf(model_data)
            if pdf_bytes:
                return {
                    "success": True,
                    "format": "pdf",
                    "size_bytes": len(pdf_bytes),
                    "generated_at": datetime.now().isoformat(),
                }
        elif format == "excel":
            excel_bytes = self.generate_excel(model_data)
            if excel_bytes:
                return {
                    "success": True,
                    "format": "xlsx",
                    "size_bytes": len(excel_bytes),
                    "generated_at": datetime.now().isoformat(),
                }

        return {
            "success": False,
            "error": "Report generation failed",
        }
